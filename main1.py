from main import DataSet
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import matplotlib.pyplot as plt
import matplotlib
from jinja2 import Environment, FileSystemLoader
import pdfkit
import openpyxl

border = Border(left=Side(border_style='thin', color='FF000000'),
                right=Side(border_style='thin', color='FF000000'),
                top=Side(border_style='thin', color='FF000000'),
                bottom=Side(border_style='thin', color='FF000000'),)

font = Font(bold=True)


class Report:

    def __init__(self, border: Border = border, font: Font = font):

        self.__border = border
        self.__headline_font = font

    def generate_excel(self, data: DataSet, salaries_by_town: {str: float}, rates_by_town: {str: float}, current_vacancy_name: str):

        wb = Workbook()
        years_sheet = wb.active
        self.__fill_year_sheet(years_sheet, data, current_vacancy_name)

        town_sheet = wb.create_sheet("Статистика по городам")
        self.__fill_town_sheet(town_sheet, salaries_by_town, rates_by_town)

        for column in ["A", "B", "C", "D", "E"]:
            years_sheet.column_dimensions[column].width =\
                max(list(map(lambda cell: len(str(cell.value)), years_sheet[column]))) + 2
            town_sheet.column_dimensions[column].width = \
                max(list(map(lambda cell: len(str(cell.value)), town_sheet[column]))) + 2

        wb.save("report.xlsx")

    def __fill_year_sheet(self, sheet, data, current_vacancy_name):

        sheet.title = "Статистика по годам"
        sheet.append(["Год", "Средняя зарплата", "Средняя зарплата - " + current_vacancy_name,
                      "Количество вакансий", "Количество вакансий - " + current_vacancy_name])

        for year in data.salaries_by_year.keys():
            sheet.append([year, data.salaries_by_year[year], data.current_salaries_by_year[year],
                          data.vacancies_count_by_year[year], data.current_count_by_year[year]])

        for row in sheet:
            for cell in row:
                if cell.row == 1:
                    cell.font = self.__headline_font
                cell.border = self.__border

    def __fill_town_sheet(self, sheet, salaries_by_town: {str: float}, rates_by_town: {str: float}):

        sheet.append(["Город", "Уровень зарплат", " ", "Город", "Доля вакансий"])

        town_rows = []
        for town_item in salaries_by_town.items():
            town_rows.append([town_item[0], town_item[1], " "])

        for i, town_item in enumerate(rates_by_town.items()):
            town_rows[i] += [town_item[0], town_item[1]]
            sheet.append(town_rows[i])

        for row in sheet:
            for cell in row:
                if cell.column == 3:
                    continue
                if cell.row == 1:
                    cell.font = self.__headline_font
                cell.border = self.__border
                if cell.column == 5:
                    cell.number_format = FORMAT_PERCENTAGE_00

    def generate_image(self, data:DataSet, salaries_by_town: {str: float},
                       rates_by_town: {str: float}, current_vacancy_name: str):

        matplotlib.rc('xtick', labelsize=8)
        matplotlib.rc('xtick', labelsize=8)
        plt.figure(figsize=(100, 100))
        fig, axs = plt.subplots(2, 2)

        self.__generate_salary_by_year_graph(data, axs,current_vacancy_name)
        self.__generate_count_by_year_graph(data, axs, current_vacancy_name)
        self.__generate_salary_by_town_graph(salaries_by_town, axs)
        self.__generate_rates_by_town_graph(rates_by_town, axs)

        fig.tight_layout()
        fig.savefig('graph.png')

    def __generate_salary_by_year_graph(self, data: DataSet, axs, current_vacancy_name: str):

        x = list(map(lambda year: int(year), data.salaries_by_year.keys()))
        axs[0, 0].bar(list(map(lambda c: c - 0.35, x)), data.salaries_by_year.values(), width=0.35, label="средняя з/п")
        axs[0, 0].bar(x, data.current_salaries_by_year.values(), width=0.35, label="з/п " + current_vacancy_name)
        axs[0, 0].set_title('Уровень зарплат по годам')
        axs[0, 0].set_xticks(x, data.salaries_by_year.keys(), rotation=90, fontsize=8)
        axs[0, 0].legend(fontsize=8)
        axs[0, 0].grid(axis='y')

    def __generate_count_by_year_graph(self, data: DataSet, axs, current_vacancy_name: str):

        x = list(map(lambda year: int(year), data.vacancies_count_by_year.keys()))
        axs[0, 1].bar(list(map(lambda c: c - 0.35, x)), data.vacancies_count_by_year.values(),
                      width=0.35, label="Количество вакансий")
        axs[0, 1].bar(x, data.current_count_by_year.values(),
                      width=0.35, label="Количество вакансий " + current_vacancy_name)
        axs[0, 1].set_title('Количество вакансий по годам')
        axs[0, 1].set_xticks(x, data.vacancies_count_by_year.keys(), rotation=90, fontsize=8)
        axs[0, 1].legend(fontsize=8)
        axs[0, 1].grid(axis='y')

    def __generate_salary_by_town_graph(self, salaries_by_town: {str: float}, axs):

        x = list(map(lambda town: town.replace(" ", "\n").replace("-", "-\n"), salaries_by_town.keys()))
        axs[1, 0].barh(x, salaries_by_town.values(), align='center')
        axs[1, 0].set_title('Уровень зарплат по городам')
        axs[1, 0].set_yticks(range(10))
        axs[1, 0].set_yticklabels(x, fontsize=6)
        axs[1, 0].grid(axis='x')
        axs[1, 0].invert_yaxis()

    def __generate_rates_by_town_graph(self, rates_by_town: {str: float}, axs):

        values = [1 - sum(rates_by_town.values())] + list(rates_by_town.values())
        labels = ["Другие"] + list(rates_by_town.keys())
        params_tuple = axs[1, 1].pie(values, labels=labels)
        axs[1, 1].set_title('Доля вакансий по городам')
        [_.set_fontsize(6) for _ in params_tuple[1]]

    def generate_pdf(self, current_vacancy_name: str, image_file: str, tables_file: str):

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        xfile = openpyxl.load_workbook(tables_file)
        years_headlines, years_values, towns_salaries_headlines,\
            towns_rates_headlines, towns_salaries_values, towns_rates_values = [], [], [], [], [], []

        years_table = xfile["Статистика по годам"]
        years_headlines = years_table[1]
        years_values = [row for row in years_table if row != years_table[1]]

        self.__fill_towns_table(xfile, towns_salaries_headlines, towns_rates_headlines,
                                towns_salaries_values, towns_rates_values)

        pdf_template = template.render({'vacancy_name': current_vacancy_name, 'image_file': image_file,
                                        'years_headlines': years_headlines, 'years_values': years_values,
                                        'towns_salaries_headlines': towns_salaries_headlines,
                                        'towns_salaries_values': towns_salaries_values,
                                        'towns_rates_headlines': towns_rates_headlines,
                                        'towns_rates_values': towns_rates_values})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        options = {'enable-local-file-access': None}
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options=options)

    def __fill_towns_table(self, xfile, towns_salaries_headlines: [str],
                           towns_rates_headlines: [str], towns_salaries_values: [float], towns_rates_values: [float]):

        town_table = xfile["Статистика по городам"]
        for row in town_table:
            salaries_value_row = []
            rates_value_row = []
            for cell in row:
                if cell.row == 1:
                    if cell.column == 1 or cell.column == 2:
                        towns_salaries_headlines.append(cell)
                    elif cell.column == 4 or cell.column == 5:
                        towns_rates_headlines.append(cell)
                else:
                    if cell.column == 1 or cell.column == 2:
                        salaries_value_row.append(cell)
                    elif cell.column == 4 or cell.column == 5:
                        rates_value_row.append(cell)
            if len(salaries_value_row) != 0:
                towns_salaries_values.append(salaries_value_row)
            if len(rates_value_row) != 0:
                towns_rates_values.append(rates_value_row)